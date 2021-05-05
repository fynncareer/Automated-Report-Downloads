# coding: utf-8
import subprocess
import argparse 
from datetime import timedelta
import psutil
import os, sys, datetime, time, argparse
from shutil import copyfile, rmtree
import openpyxl
import pandas as pd
import numpy as np
import logging 
import csv

import smtplib 
from email.mime.text import MIMEText

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from selenium.webdriver.common.action_chains import ActionChains

today = datetime.date.today()
now = datetime.datetime.now() 
year = now.year 
WW = today - datetime.timedelta(days = 7)
WW = WW.strftime("%U")
YYYYWW = '{}{}'.format(year, WW)

try: 
	if os.path.exists("C:\\Users\user\AppData\Local\Temp"):
		shutil.rmtree("C:\\Users\user\AppData\Local\Temp") # remove temp files because they can interfere with Selenium 
except BaseException:
	pass 


def DeleteOldFiles(path, days):
	now = time.time()
	try:
		for f in os.listdir(path):
			f = os.path.join(path, f)
			if os.stat(f).st_mtime < now - days * 86400:
				if os.path.isfile(f):
					os.remove(f)
					files = os.listdir(path)
	except BaseException as error:
		pass 

				
def LoggingConfig(job, level): 
	DeleteOldFiles("C:\\Logs", 14)
	
	fname = job + " {}".format(datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S'))
	filename = 'C:\\Logs\\%s.txt' % fname

	config = logging.basicConfig(filename = filename, level = level, format = '%(levelname)s:%(asctime)s:%(message)s') # filemode, 'w'
	return config, filename	
	
def PrintLog(statement):
	logging.info(statement)
	print(statement) 
	
	
def DaysFromMonday(format, days_to_subtract):
	# weekday values:
	# Mon = 0, Tues = 1, Weds = 2, Thu = 3, Fri = 4, Sat = 5, Sun = 6
	
	parser = argparse.ArgumentParser()
	parser.add_argument("--end_date") 
	args = parser.parse_args()
	end_date = args.end_date
	
	if end_date:
		end_date_obj = datetime.datetime.strptime(end_date,'%d/%m/%Y').date()
		end_date_str = datetime.datetime.strftime(end_date_obj, format)
		
		start_date_obj = end_date_obj - datetime.timedelta(days = 6)
		start_date_str = datetime.datetime.strftime(start_date_obj, format)
		
	else: 
		today = datetime.date.today()
		now = datetime.datetime.now() 
		
		weekday = today.weekday()
		weekday_day = days_to_subtract 
		
		if 0 <= weekday <= 6:
			weekday_day = weekday_day + weekday 

		end_date_obj = today - datetime.timedelta(days = weekday_day)
		end_date_str = datetime.datetime.strftime(end_date_obj, format)
		start_date_obj = end_date_obj - datetime.timedelta(days = 6)
		start_date_str = datetime.datetime.strftime(start_date_obj, format)
		
	
	print("Start Date: ", start_date_str)
	print("End Date: ", end_date_str)
	return start_date_str, start_date_obj, end_date_str, end_date_obj
	

def GetLogin(job):
	wb = openpyxl.load_workbook("C:\\Scripts\\Config.xlsx")
	sheet1 = wb.get_sheet_by_name('Sheet1')
	
	spreadsheet_dict = {
	"Amazon Digital Daily Pre-Orders and NYP": 2,
	"Amazon Digital Weekly Sales": 3,
	"Google": 4, 
	"ASDA": 5,
	"Amazon Physical": 6,
	"Apple Ebooks": 7,
	"Apple iTunes": 8,
	"Nielsen Bookscan Key Account": 9,
	"Amazon Daily Digital Orders": 10,
	"Amazon All Geo Data": 11
	}
	
	username = sheet1.cell(row = spreadsheet_dict[job], column = 2).value
	password = sheet1.cell(row = spreadsheet_dict[job], column = 3).value
	
	return username, password 
	
	
def RenameDownloadedFile(src, dst, wait_time):
	time.sleep(wait_time)

	PrintLog("Downloading File.")
	
	os.rename(src, dst)

	time.sleep(wait_time)

	PrintLog("File renamed to: " + os.path.basename(dst))
	
# def SpreadSheet_RenameSheet(file, new_sheetname):
	# ss = openpyxl.load_workbook(file)
	# #printing the sheet names
	# sheet = ss.get_sheet_names()
	# sheet = sheet[0]
	
	# ss_sheet = ss.get_sheet_by_name(sheet)
	
	# ss_sheet.title = new_sheetname
	# ss.save(file)

# def FileBackup(*Files):
	# # create a backup of this weeks files before making any changes to them.
	# for File in Files:
		# File_1, ext = os.path.splitext(File)
		# Backup_File = File_1 + '_{}'.format(today_date) + ext
		# print Backup_File
		# return Backup_File
		
		# #copyfile(File, Backup_File)
		
def FilesToDownload(download_dir, report_name, start_date_dateobj, end_date_dateobj):
	DownloadList = []
	for process_date in daterange(start_date_dateobj, end_date_dateobj):
		#report_date = process_date.strftime('%Y%m%d')
		newfilename = os.path.join(download_dir, report_name).format(process_date.strftime('%Y%m%d'))
		DownloadList.append(newfilename)
	
	return DownloadList		
	
def CheckFilesExist(DownloadList):
	SecondList = []
	PrintLog("This weeks files:")
	for x in DownloadList:
		PrintLog(os.path.basename(x))
		if os.path.exists(x) == False:
			#print "{}".format(os.path.basename(x))
			SecondList.append(x)
		else:
			PrintLog("{} already exists".format(os.path.basename(x)))
	return SecondList 
		
def TransferFile(Source, Target, job = None):
	time.sleep(10)
	
	just_filename = os.path.basename(Source)
	Target = os.path.join(Target, just_filename)
	
	folder_list = ["01 Daily Pre-Orders and NYP - Amazon Digital",
	"02 Weekly Sales - Amazon Digital", 
	"03 Google",
	"04 Asda", 
	"05 Daily All Geographic Data - Amazon Physical", 
	"06 Daily Pre-Orders and NYP - Amazon Physical", 
	"07 Amazon Weekly Search Terms", 
	"08 Alternative Purchase - Amazon Physical", 
	"09 Customer Purchase Behaviour - Amazon Physical",
	"10 Item Viewing Pre-Purchase - Amazon Physical", 
	"11 Market Basket Analysis - Amazon Physical", 
	"12 Apple Books",
	"13 Apple iTunes",
	"14 Bookscan Key Account Data"
	]	

	weekday = today.weekday()
	weekday_day = 8 
	
	if 0 <= weekday <= 6:
		weekday_day = weekday_day + weekday 

	sunday_before_last = today - datetime.timedelta(days = weekday_day)
	sunday_before_last = datetime.datetime.strftime(sunday_before_last, '%Y.%m.%d')
	
	try:
		copyfile(Source, Target)
	except OSError:
		print("File already exists in target location.")
	
	if job: 
		uk_folder = job 
		
		uk_server = "\\\\ukserver\\source\Week %s wc %s" % (WW, sunday_before_last)
		uk_folder = "\\\\ukserver\\UK folder\\Week %s wc %s\\%s" % (WW, sunday_before_last, uk_folder)
		uk_folder = os.path.join(uk_folder, just_filename)
		
		if os.path.exists(uk_server) == False:
			try: 
				os.makedirs(uk_server)
				os.chdir(uk_server)
				for folders in uk_folder_list:
					os.makedirs(folders)
			except OSError:
				pass 
		try:
			copyfile(Source, uk_folder)	 
			print_file_transfer_a = """
			****** File Transfer ******
			Target: %s
			Backup: %s
			****** File Transfer ******
			""" % (Target, uk_folder)
			PrintLog(print_file_transfer_a)
		except BaseException as e:
			pass 
			

		
	else:
		print_file_transfer_b = """
		****** File Transfer ******
		Target: %s
		****** File Transfer ******
		""" % (Target)
		PrintLog(print_file_transfer_b)

		
 		
def SendEmail(recipient, job, log):
	#msg:
	#	JOB RUN: [name of job], datetimestamp 
	#	DURATION: [time log of script]
	#	STATUS: [FAILED/SUCCEEDED]
	#	MESSAGES: [LB + Indent - Log of files downloaded here - want either list of downloaded files OR error]
	
	log_contents = open(log, "r")
	msg = MIMEText(log_contents.read())
	log_contents.close()
	
	sender = 'service account'
		
	msg['Subject'] = "DATA WAREHOUSE {}".format(job) 
	msg['From'] = 'service account'
	msg['To'] = recipient

	s = smtplib.SMTP('relay-co.uk.randomhouse.com') 
	s.sendmail(sender, recipient.split(','), msg.as_string())
	s.quit()		

########## SELENIUM FUNCTIONS

## method under browser Object? 
def setdate(element, newdate, field):
	for i in range(11):
		element.send_keys(Keys.BACK_SPACE)
		time.sleep(0.1)
	element.send_keys(newdate)
	time.sleep(0.1)
	# element.send_keys(newdate) doesn't always send all keys. The send_keys command is flaky, possibly a defect with selenium. The code below will check that the date has been set correctly, and if not then retry sending it. 
	if element.get_attribute('value') == newdate:
		print("Entered %s in %s field" % (newdate, field))
		return True
	else:
		print("FAILED to enter %s in %s field" % (newdate, field))
		print("Re-attempting...")
		for i in range(0,5):
			element.send_keys(newdate)
			if element.get_attribute('value') == newdate:
				print("Entered %s in %s field" % (newdate, field))
				break
		else:
			return 
			print("Failed to setdate(enter_prompt_date) in From/To Date box after 5 attempts.")
			print("Terminating script.")
			sys.exit()

#test this? 
def daterange(fromdate, todate):
    ldate = fromdate
    while ldate <= todate:
        yield ldate
        ldate += timedelta(days=1)	

# should go in apple_scripts not here.
def AppleSignIn(driver, username, password):
		# deals with ocassional login error on Apple site ("Failed to verify your identity"). Not sure why it's happening, so re-attempt login.
		for i in range(0,5):
			appleId = driver.find_element_by_xpath('') # 
			appleId.send_keys(username)
			time.sleep(2)
			pwd = driver.find_element_by_xpath('') # //*[@id="pwd"]
			pwd.send_keys(password)
			time.sleep(2)
			signin = driver.find_element_by_xpath("")
			signin.click()
			
			time.sleep(10)
			
			if driver.current_url == "https://itunesconnect.apple.com/logout":
				pass
			else:
				break
		
		return driver 


###############		

# Browser
def CloseOpenBrowserWindows(Browser):
	process_name = Browser + ".exe" 
	for proc in psutil.process_iter(): 
		try:
			process = psutil.Process(proc.pid)# Get the process info using PID
			pname = process.name()
			if pname == process_name: 
				proc.kill()
		except(psutil.NoSuchProcess, psutil.AccessDenied) as e:
			print(e, "closing open browser windows failed")
			break


def StartFireFox(base_url, download_dir, wait_time):
	firefoxprofile = r'C:\Scripts\firefoxprofile'
	CloseOpenBrowserWindows("firefox")

	fp = webdriver.FirefoxProfile(firefoxprofile) 
	fp.set_preference("browser.download.folderList", 2)
	fp.set_preference("browser.download.manager.showWhenStarting", False)
	fp.set_preference("browser.download.dir", download_dir)
	fp.set_preference("browser.helperApps.alwaysAsk.force", False)
	fp.set_preference("browser.helperApps.neverAsk.openFile", "text/csv, text/css, application/vnd.ms-excel,application/a-gzip;charset=utf-8, application/json;charset=UTF-8, application/gzip, application/x-gzip, application/zip")
	fp.set_preference("browser.helperApps.neverAsk.saveToDisk","text/plain, text/css, application/json;charset=UTF-8,application/a-gzip;charset=utf-8, application/zip, application/gzip, application/vnd.ms-excel, text/csv, application/csv, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, text/comma-separated-values, application/download, application/octet-stream, binary/octet-stream, application/binary, application/x-unknown")
	fp.update_preferences()
	
	driver = webdriver.Firefox(firefox_profile = fp)
	driver.implicitly_wait(wait_time) 
	driver.get(base_url)
	
	return driver 


# Browser	
def StartChrome(base_url, download_dir, wait_time):
	# remove files more than 2 weeks old in download_dir 
	DeleteOldFiles(download_dir, 14)
				
	# start the chrome driver
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_experimental_option('prefs', {'download.default_directory': download_dir})

	driver = webdriver.Chrome(executable_path = "C:\Program Files (x86)\Google\Chrome\chromedriver.exe", chrome_options = chrome_options)
	driver.implicitly_wait(wait_time)
	driver.get(base_url) 
	
	return driver 


if __name__ == "__main__":
	pass
else:
	try: 
		if os.path.exists("C:\\Users\user\AppData\Local\Temp"):
			shutil.rmtree("C:\\Users\user\AppData\Local\Temp")
	except BaseException:
		pass 
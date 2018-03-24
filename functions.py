"""
	- Each script can be run individually.
		- There is no harm in running the same script multiple times a day. 
	- Dates are generated with a default date.
		- To use a different date, run the script with command line parameter --start_date [dd/mm/yyyy]  
	- Login details, file names, urls, download and target locations are stored in spreadsheet: config.xlsx 
"""

#### CONFIG
server = server_here
log_path = log_path_here
firefoxprofile = firefox_profile_here 
file_config = config_file_here 
#### CONFIG

import argparse 
from datetime import timedelta
import psutil
import os, sys, datetime, time, argparse
import openpyxl
import pandas as pd
import logging 
import textwrap 

import smtplib 
from email.mime.text import MIMEText

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, InvalidElementStateException, TimeoutException, StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import pandas as pd 
from openpyxl import load_workbook
import shutil 

today = datetime.date.today()
now = datetime.datetime.now() 
month_year_current = datetime.datetime.strftime(today, "%B %Y")
now_timestamp = datetime.datetime.strftime(datetime.datetime.now(), "%d/%m/%Y %H:%M")
year = now.year 
	
def LogError(error): 
	
	filename = "{}.txt".format(os.path.join(log_path, sys.argv[0][:-3]))
		
	try:
		if os.path.getsize(filename) > 3000000: #3MB
			with open(filename, 'w'): # truncate the file.
				pass  
	except OSError:
		pass 
	
	logging.basicConfig(level = logging.DEBUG, filename = filename, format = '%(asctime)s:%(message)s')
	logger = logging.getLogger(__name__)
	logger.exception(error) 	

def GetFiles(script_name):
	try:
		wb = openpyxl.load_workbook(file_config) 
	except PermissionError as e:
		sys.exit(e)
		
	files = wb.get_sheet_by_name("FILES")

	for row in files:
		for cell in row:
			if cell.value == script_name: 
				row_val = cell.row 
		
	job = files.cell(row = row_val, column = 2).value
	login = files.cell(row = row_val, column = 3).value
	download_location = files.cell(row = row_val, column = 4).value
	download_file = files.cell(row = row_val, column = 5).value
	target_location = files.cell(row = row_val, column = 6).value
	target_file = files.cell(row = row_val, column = 7).value
	base_url = files.cell(row = row_val, column = 8).value
	
	return job, login, download_location, download_file, target_location, target_file, base_url

def GetLogin(job):
	try:
		wb = openpyxl.load_workbook(file_config)
	except PermissionError as e:
		sys.exit(e)
	
	logins = wb.get_sheet_by_name('LOGINS')
	
	for row in logins:
		for cell in row:
			if cell.value == job: 
				row_val = cell.row 
	
	username = logins.cell(row = row_val, column = 2).value
	password = logins.cell(row = row_val, column = 3).value
	
	return username, password 
	
def Dates(format, script): 
	
	parser = argparse.ArgumentParser()
	parser.add_argument("--start_date")
	parser.add_argument("--end_date")
	args = parser.parse_args()
	start_date = args.start_date
	end_date = args.end_date
	
	if end_date:	
		if (datetime.datetime.strptime(start_date, "%d/%m/%Y") >= datetime.datetime.strptime(end_date, "%d/%m/%Y")) is True:
			sys.exit("End Date cannot be before Start Date")
			
	if start_date:		
		if (datetime.datetime.strptime(start_date, "%d/%m/%Y") >= datetime.datetime.today()) is True:
			sys.exit("Start Date cannot be today or in the future.")	
		
		start_date_obj = datetime.datetime.strptime(start_date,'%d/%m/%Y').date()
		start_date = datetime.datetime.strftime(start_date_obj, format)
	else:
		today = datetime.date.today()
		now = datetime.datetime.now() 
		
		weekday = today.weekday()	# Mon = 0, Tues = 1, Weds = 2, Thu = 3, Fri = 4, Sat = 5, Sun = 6 
		
		if script in ("SaleDiag_W_OrdRev_Phy.py", "SaleDiag_W_ShipCOGS_Phy.py", "SaleDiag_W_ShipRev_Phy.py", "FcstInv_W_OrdUnit_Phy.py"):			
			weekday_day = 8 
			if 0 <= weekday <= 6:
				weekday_day = weekday_day + weekday 
				
		elif script in ("Alternative_Purchase_W_Phy.py", "Market_Basket_Analysis_W_Phy.py", "Repeat_Purchase_W_Phy.py"):			
			weekday_day = 2
	
			if 0 <= weekday <= 6:
				weekday_day = weekday_day + weekday
				
		else:
			weekday_day = 1
		
		start_date_obj = today = datetime.date.today() - datetime.timedelta(days = weekday_day) # weekday_day 
		start_date = datetime.datetime.strftime(start_date_obj, format)
		
	if end_date:
		end_date_obj = datetime.datetime.strptime(end_date,'%d/%m/%Y').date()
		end_date = datetime.datetime.strftime(end_date_obj, format)
		
	print(start_date)
	if end_date:
		print(end_date) 
	
	return start_date, end_date
	
def DateRange(start_date,end_date):
	start_date = datetime.datetime.strptime(start_date, "%Y%m%d").date()
	if end_date is None:
		end_date = start_date
	else:
		end_date = datetime.datetime.strptime(end_date, "%Y%m%d").date()
	print(start_date, end_date)
	
	while start_date <= end_date:
		yield start_date
		start_date += timedelta(days=1)
		
def checkEndDateFilesExists(start_date, end_date, target_location, target_file_original, source):
	for x in DateRange(start_date, end_date):
		start_date_b = datetime.datetime.strftime(x, "%Y%m%d")
		target_file_b = os.path.join(target_location, target_file_original.format(source[:2], start_date_b))
		
		filesToDownload = []
		if os.path.exists(target_file_b) == False:
			filesToDownload.append(target_file_b)
			
		if filesToDownload:
			return True 
	
def RenameDownloadedFile(src, dst, wait_time):
	count = 10
	wait_time = wait_time / count 
	for i in range(count):
		try:
			with open(src, 'rb') as _:
				break
		except IOError:
			time.sleep(wait_time)
	else:
		raise IOError('Could not access {} after {} attempts.'.format(src, count))
	
	try:
		os.rename(src, dst)
	except FileExistsError:
		os.remove(dst)
		os.rename(src, dst)
		
	print("Src file downloaded and renamed to: " + os.path.basename(dst))	

			
def SendEmail(recipient, sender, job, duration, status, message):
	msg_str = textwrap.dedent("""
	JOB RUN: Python job "{}" completed on {}.  
	DURATION: The job took {} seconds to run.
	STATUS: {}
	MESSAGES: 
	{}  
	""".format(job, now_timestamp, duration, status, message))
	# messages print log file path 
	msg = MIMEText(msg_str)

	msg['Subject'] = '{}: "{}" run on {}'.format(status.capitalize(), job, server) # caps (status{ {job_run}
	msg['From'] = sender
	msg['To'] = recipient

	s = smtplib.SMTP(smtp_adress_here) 
	s.sendmail(sender, recipient.split(','), msg.as_string())
	s.quit()
	
def StringReplace(target_file, str_src, str_dest):
	#string replace any files in the spreadsheet 

	time.sleep(5)
	data = pd.read_excel(target_file, sheet_name = None)
	book = load_workbook(target_file)
	writer = pd.ExcelWriter(target_file, engine='openpyxl') 
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

	for sheet in data:
		print(sheet)
		data[sheet].replace(str_src, str_dest).to_excel(writer, index = False, sheet_name = sheet) # target_file[:-5] + "_test.xlsx"
		#data.to_excel(writer, "Main")
		writer.save()

########## SELENIUM FUNCTIONS
def SiteLogin(driver, username, password):
	time.sleep(3)
	driver.find_element_by_id("email").clear()
	driver.find_element_by_id("email").send_keys(username)
	driver.find_element_by_id("password").clear()
	driver.find_element_by_id("password").send_keys(password, Keys.ENTER)
	
	time.sleep(3)
	try:
		driver.find_element_by_class_name("a-button-input").click()
	except NoSuchElementException:
		pass
		
	return driver

def SelectSalesView(driver, view):
	if view == "Ordered Revenue":
		pass # "Ordered Revenue" is the default dropdown option.  
	else:
		driver.find_element_by_css_selector("").click()
		driver.find_element_by_link_text(view).send_keys(Keys.ENTER)
	
	return driver

def SelectReportingPeriod(driver, job, reporting_period):
	# Options: Daily, Weekly, Monthly, Quarterly, Yearly
	if reporting_period == "Weekly":
		pass # "Weekly" is the default dropdown option.	
	else:
		css_selector = {"Sales Diagnostic": 3, "Pre-Orders": 2, "GeoSales": 4, "Customer Purchase Behaviour": 3}
		css_selector = "".format(css_selector[job])
		driver.find_element_by_css_selector(css_selector).send_keys(Keys.ENTER) #Dropdown	
		driver.find_element_by_link_text(reporting_period).send_keys(Keys.ENTER) #Option 
			
	return driver
	
def SelectDate(driver, from_date, to_date = False):	
	from_date_day = datetime.datetime.strptime(from_date, "%Y%m%d").strftime("%#d")
	from_date_month_year = datetime.datetime.strptime(from_date, "%Y%m%d").strftime("%B %Y")
	
	xpath_date = "//div[@aria-label='day-{}']".format(from_date_day)
	
		# if xpath_date = 27 and week 0 contains xpath_date, click next value. 
		
	def DatePicker(css_selector):
		try:
			driver.find_element_by_css_selector(css_selector).click() 
		except (NoSuchElementException, InvalidElementStateException):
			driver.find_element_by_css_selector(".date-picker-clear").click()
			driver.find_element_by_css_selector(css_selector).click()
		
		# Date Picker FROM DATE
		datepicker_month_xpath = ""

		if (from_date_month_year == month_year_current) is not True:		#current month 
			datepicker_month = driver.find_element_by_css_selector(datepicker_month_xpath).text # November 2017
			while datepicker_month != from_date_month_year:
				driver.find_element_by_css_selector("a.react-datepicker__navigation:nth-child(2)").click() # navigate to previous month		
				time.sleep(3)
				datepicker_month = driver.find_element_by_css_selector(datepicker_month_xpath).text
				
			time.sleep(2)
			week0 = driver.find_element_by_css_selector("") #week0 is the first line (week) of the calendar. If "29" is in the first week, you don't want to click that value, but the second one. Hence the IF statement below. 
			
			if from_date_day in week0.text and from_date_day in ('24', '25', '26', '27', '28', '29', '30', '31'):
				driver.find_elements_by_xpath(xpath_date)[1].click()
			else:
				driver.find_elements_by_xpath(xpath_date)[0].click()
				
		else:
			try:
				driver.find_element_by_xpath(xpath_date).click() 
			except (NoSuchElementException, InvalidElementStateException) as e:
				driver.find_element_by_css_selector(css_selector).click()
				driver.find_element_by_xpath(xpath_date).click() 
	#daily  
	DatePicker("") # from date
	if (datetime.datetime.strftime(today - datetime.timedelta(days = 1), "%Y%m%d") != from_date): # check we're not using most recent date. If using most recent date then no need for datepicker(to_date).
		try:
			DatePicker("") # to date 
		except (NoSuchElementException, InvalidElementStateException):
			pass 
	try:	
		driver.find_element_by_css_selector("").click()  # apply button 
	except NoSuchElementException:
		pass

	return driver 

def SelectColumns(driver, job, columns):
	if job == "Sales Diagnostic": 
		css_selector = "" # Sales Diagnostic 
		xpath_options = ""
	elif job in ("Pre-Orders", "Forecast and Inventory Planning", "Customer Purchase Behaviour"):
		css_selector = "" # Pre-Orders 
		xpath_options = ""
	elif job in ("GeoSales"):
		css_selector = ""

	xpath_dropdown = ""

	time.sleep(2)
			
	for col in columns:	
		driver.find_element_by_css_selector(css_selector).send_keys(Keys.ENTER)
		driver.find_element_by_link_text(col).send_keys(Keys.ENTER)
		
	time.sleep(5) 
	
	return driver 
	
def DownloadFile(driver, job):
	time.sleep(5)
	dropdown = driver.find_element_by_css_selector("")
	dropdown.send_keys(Keys.ENTER)
	
	time.sleep(10)
	if job in ("Sales Diagnostic", "GeoSales"):
		csv_button = driver.find_elements_by_link_text("As Excel Workbook (.xlsx)")
		csv_button = driver.find_element_by_xpath("")
		csv_button.send_keys(Keys.ENTER) # 0: All Views. 1: Summary View, 2: Detail View 
	elif job in ("Pre-Orders", "Forecast and Inventory Planning", "Customer Purchase Behaviour"):
		driver.find_element_by_xpath("").send_keys(Keys.ENTER) 

	return driver 		
	
	
#####BROWSER

def CloseOpenBrowserWindows(Browser):
	process_name = Browser + ".exe" 
	for proc in psutil.process_iter(): 
		try:
			process = psutil.Process(proc.pid)# Get the process info using PID
			pname = process.name()
			if pname == process_name: 
				proc.kill()
		except(psutil.NoSuchProcess, psutil.AccessDenied) as e:
			break


def StartFireFox(base_url, download_dir, wait_time):
	
	CloseOpenBrowserWindows("firefox")

	fp = webdriver.FirefoxProfile(firefoxprofile) 
	fp.set_preference("browser.download.folderList", 2)
	fp.set_preference("browser.download.manager.showWhenStarting", False)
	fp.set_preference("browser.download.dir", download_dir)
	fp.set_preference("security.sandbox.content.level", 5)
	fp.set_preference("browser.helperApps.alwaysAsk.force", False)
	fp.set_preference("browser.helperApps.neverAsk.saveToDisk","text/plain, application/zip, application/vnd.ms-excel, text/csv, application/csv, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, text/comma-separated-values, application/download, application/octet-stream, binary/octet-stream, application/binary, application/x-unknown")
	fp.update_preferences()
	
	driver = webdriver.Firefox(firefox_profile = fp)
	driver.implicitly_wait(wait_time) 
	driver.get(base_url)
	
	return driver 

def DeleteOldFiles(path, days): # delete old files from the target file 
	now = time.time()
	try:
		for f in os.listdir(path):
			f = os.path.join(path, f)
			if os.stat(f).st_mtime < now - days * 86400:
				if os.path.isfile(f):
					os.remove(f)
					files = os.listdir(path)
	except BaseException as error:
		pass 

if __name__ == "__main__":	
	pass 
else:
	print("Running script {}".format(sys.argv[0]))
	CloseOpenBrowserWindows("firefox")	
	DeleteOldFiles(target_location_folder_here, 7) 